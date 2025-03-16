import openpyxl

wb = openpyxl.load_workbook('productSales.xlsx')
sheet = wb['Sheet']

PRICE_UPDATES = {
    'Garlic': 3.07,
    'Celery': 1.19,
    'Lemon': 1.27
}

for rowNum in range(2, sheet.max_row):
    produceNAme = sheet.cell(row=rowNum, column=1).value
    if produceNAme in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceNAme]
wb.save('updatedProduceSales.xlsx')